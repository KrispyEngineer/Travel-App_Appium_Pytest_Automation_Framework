import inspect
import logging
import openpyxl

# This is decommisioned

class BaseClass:


    @staticmethod
    def get_dataload():
        workbook = openpyxl.load_workbook("..//EXCEL//testdata.xlsx")
        sheet = workbook.active
        Dict = {}
        store = []
        for i in range(2, sheet.max_row+1):

            for j in range(1,sheet.max_column+1):

                Dict[sheet.cell(row=1,column=j).value] = sheet.cell(row=i,column=j).value
            store.append(Dict)
        # print(sheet.max_row)
        # print(sheet.max_column)
        return store


