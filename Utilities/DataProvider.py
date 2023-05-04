import openpyxl


def get_dataload(sheetName):
    workbook = openpyxl.load_workbook("..//EXCEL//testdata.xlsx")
    sheet = workbook[str(sheetName)]
    finaldata = []

    for i in range(2, sheet.max_row + 1):
        data = []
        for j in range(1, sheet.max_column + 1):

            data.insert(j-1,sheet.cell(row=i, column=j).value)
        finaldata.insert(i-2,data)
    #print(sheet.max_row)
    #print(sheet.max_column)

    return finaldata

print(get_dataload("TestHotel"))