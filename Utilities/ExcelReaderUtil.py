import openpyxl


def get_MaxRows(path,sheetName):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetName]
    return sheet.max_row

def get_MaxCol(path,sheetName):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetName]
    return sheet.max_cols

def get_CellData(path,sheetName,rowNum,colNum):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetName]
    return sheet.cell(row=rowNum, column=colNum).value

def get_SetCellData(path,sheetName,rowNum,colNum,data):
    workbook = openpyxl.load_workbook(path)
    sheet = workbook[sheetName]
    sheet.cell(row=rowNum, column=colNum).value = data

print(get_MaxRows("..//EXCEL//testdata.xlsx",'Sheet1'))
print(get_MaxCol("..//EXCEL//testdata.xlsx",'Sheet1'))